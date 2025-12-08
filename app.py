
# app.py
import os
import threading
from datetime import datetime

from flask import Flask, jsonify, request, abort
import pandas as pd
from openpyxl import load_workbook

app = Flask(__name__)

# -------- Configuration --------
# On Render, set EXCEL_PATH to a path on an attached Disk (e.g., /data/Data.xlsx)
EXCEL_PATH = os.getenv("EXCEL_PATH", "/data/Data.xlsx")
# Optional: protect write ops
API_KEY = os.getenv("API_KEY", "")  # set in Render if you want to secure /update and /reload

# -------- Shared state --------
final_df = pd.DataFrame()
_df_lock = threading.RLock()
_last_loaded_at = None


def _normalize(s):
    """Lowercase + strip helper for consistent comparisons."""
    return str(s).strip().lower() if s is not None else ""


def _to_json_safe(df: pd.DataFrame):
    """Replace NaN with None for clean JSON."""
    return df.where(pd.notnull(df), None)


def _require_api_key():
    """Require API key if configured."""
    if API_KEY:
        provided = request.headers.get("X-API-Key")
        if provided != API_KEY:
            abort(401, description="Unauthorized")


def load_excel():
    """
    Load ALL sheets into one DataFrame (adds SOURCE_SHEET).
    Normalizes datetime columns to string for stable JSON.
    """
    global final_df, _last_loaded_at

    # Read all worksheets at once
    xls = pd.read_excel(EXCEL_PATH, sheet_name=None, engine="openpyxl")

    combined = []
    for sheet_name, df in xls.items():
        df["SOURCE_SHEET"] = sheet_name
        combined.append(df)

    if not combined:
        raise FileNotFoundError(f"No sheets found at path: {EXCEL_PATH}")

    df_all = pd.concat(combined, ignore_index=True)

    # Convert datetime columns to strings for JSON stability
    for col in df_all.columns:
        if pd.api.types.is_datetime64_any_dtype(df_all[col]):
            df_all[col] = df_all[col].dt.strftime("%Y-%m-%d %H:%M:%S")

    with _df_lock:
        final_df = df_all
        _last_loaded_at = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
    app.logger.info(f"Excel loaded: {len(df_all)} rows from {len(combined)} sheet(s).")


@app.route("/")
def home():
    return "✅ API is running! Use /data to get data or /update to update Excel."


@app.route("/health")
def health():
    with _df_lock:
        rows = int(final_df.shape[0]) if not final_df.empty else 0
    return jsonify({
        "status": "ok",
        "excel_path": EXCEL_PATH,
        "last_loaded_at": _last_loaded_at,
        "rows": rows
    }), 200


@app.route("/reload", methods=["POST"])
def reload_endpoint():
    """Manually reload the Excel into memory (useful after external changes)."""
    _require_api_key()
    try:
        load_excel()
        return jsonify({"status": "success", "message": "Excel reloaded"}), 200
    except Exception as e:
        app.logger.exception("Reload failed")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/data", methods=["GET"])
def get_all_data():
    """
    Optional query params:
      - sheet: filter by SOURCE_SHEET (case-insensitive)
      - client_code: exact match on 'CLIENT CODE' (case-insensitive)
      - limit, offset: pagination
      - columns: comma-separated selection of columns to return
    """
    sheet = request.args.get("sheet")
    client_code = request.args.get("client_code")
    limit = int(request.args.get("limit", "0") or 0)
    offset = int(request.args.get("offset", "0") or 0)
    columns_param = request.args.get("columns")

    with _df_lock:
        if final_df is None or final_df.empty:
            return jsonify({"status": "error", "message": "No data loaded"}), 503

        df = final_df

        if sheet:
            df = df[df["SOURCE_SHEET"].str.lower() == sheet.strip().lower()]

        if client_code:
            # Locate 'CLIENT CODE' dynamically (case-insensitive match)
            client_col = None
            for col in df.columns:
                if _normalize(col) == "client code":
                    client_col = col
                    break
            if not client_col:
                return jsonify({"status": "error", "message": "CLIENT CODE column not found"}), 400
            df = df[df[client_col].apply(_normalize) == _normalize(client_code)]

        if offset:
            df = df.iloc[offset:]
        if limit:
            df = df.iloc[:limit]

        if columns_param:
            wanted = [c.strip() for c in columns_param.split(",") if c.strip()]
            existing = [c for c in wanted if c in df.columns]
            if existing:
                df = df[existing]

        data = _to_json_safe(df).to_dict(orient="records")
    return jsonify(data), 200


@app.route("/update", methods=["POST"])
def update_excel():
    """
    Body JSON example:
    {
      "sheet": "corp",
      "client_code": "ACME123",
      "column": "STATUS",
      "new_value": "Cross-Sell"
    }

    - Finds row by CLIENT CODE (case-insensitive).
    - Updates the target column (case-insensitive header match).
    - Updates/creates STATUS_UPDATED_AT (uppercase) with current timestamp.
    - Saves to EXCEL_PATH and reloads in-memory DataFrame.
    """
    _require_api_key()
    payload = request.get_json(silent=True) or {}
    sheet = payload.get("sheet")
    client_code = payload.get("client_code")
    column = payload.get("column")
    new_value = payload.get("new_value")

    if not all([sheet, client_code, column]):
        return jsonify({"status": "error",
                        "message": "Missing one of: sheet, client_code, column"}), 400

    try:
        with _df_lock:
            wb = load_workbook(EXCEL_PATH)
            if sheet not in wb.sheetnames:
                return jsonify({"status": "error", "message": f"Sheet '{sheet}' not found"}), 404

            ws = wb[sheet]

            # Build case-insensitive header map (zero-based)
            headers = [cell.value for cell in ws[1]]
            header_map = {_normalize(h): idx for idx, h in enumerate(headers)}

            # Locate CLIENT CODE
            if "client code" not in header_map:
                return jsonify({"status": "error", "message": "CLIENT CODE column not found"}), 400
            client_col_idx = header_map["client code"]

            # Locate target column
            target_key = _normalize(column)
            if target_key not in header_map:
                return jsonify({"status": "error", "message": f"Column '{column}' not found"}), 400
            target_col_idx = header_map[target_key]

            # Ensure STATUS_UPDATED_AT exists (store the new index explicitly)
            ts_key = "status_updated_at"
            if ts_key not in header_map:
                new_col = ws.max_column + 1  # 1-based for openpyxl
                ws.cell(row=1, column=new_col, value="STATUS_UPDATED_AT")
                timestamp_col_idx = new_col - 1  # convert to zero-based for row[] access below
            else:
                timestamp_col_idx = header_map[ts_key]

            # Find row by client code
            found = False
            for row in ws.iter_rows(min_row=2):
                cell_val = row[client_col_idx].value
                if _normalize(cell_val) == _normalize(client_code):
                    # Update value
                    row[target_col_idx].value = new_value
                    # Update timestamp now (1-based write through row[] index)
                    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    row[timestamp_col_idx].value = current_time
                    found = True
                    break

            if not found:
                return jsonify({"status": "error", "message": "Client Code not found"}), 404

            # Save workbook
            try:
                wb.save(EXCEL_PATH)
            except PermissionError:
                return jsonify({"status": "error",
                                "message": "Excel file is locked/open. Close it and try again."}), 423

        # Reload memory cache after write
        load_excel()

        return jsonify({
            "status": "success",
            "message": f"Updated '{column}' for '{client_code}' on '{sheet}'"
        }), 200

    except Exception as e:
        app.logger.exception("Update failed")
        return jsonify({"status": "error", "message": str(e)}), 500


if __name__ == "__main__":
    # Local dev
    try:
        load_excel()
    except Exception as e:
        app.logger.error(f"Initial load failed: {e}")
