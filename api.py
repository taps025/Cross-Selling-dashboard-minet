
# api.py
from flask import Flask, jsonify, request
import pandas as pd
from openpyxl import load_workbook
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import threading
from datetime import datetime
import time
import traceback
import os
import re
import sqlite3
import logging
from typing import List, Dict, Any, Optional

# -----------------------------
# Config
# -----------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Point these to your persistent disk paths in Render (e.g., /data/Data.xlsx, /data/overrides.db)
EXCEL_FILE = os.environ.get("EXCEL_FILE", os.path.join(BASE_DIR, "Data.xlsx"))
SHEETS = [s.strip() for s in os.environ.get("SHEETS", "corp,EB,SS,PLD,AFFINITY,MINING").split(",") if s.strip()]
OVERRIDES_DB = os.environ.get("OVERRIDES_DB", os.path.join(BASE_DIR, "overrides.db"))

# Allowed values for the Status column (optional strictness)
ALLOWED_STATUS_VALUES = {"CROSS-SELL", "SHARED CLIENT"}

# Flask app
app = Flask(__name__)

# Configure logging (Render captures stdout/stderr)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s"
)

# In-memory cache of combined data
final_df = pd.DataFrame()

# One-time initialization guard (works for Gunicorn + local)
_init_lock = threading.Lock()
_initialized = False


# -----------------------------
# Helpers: canonicalization & header/row lookup
# -----------------------------
def canon(s: str) -> str:
    """Canonicalize header names to avoid punctuation/case mismatches."""
    if s is None:
        return ""
    s = re.sub(r"[`.,:\-\[\]]+", "", str(s))  # strip punctuation variants
    s = re.sub(r"\s+", " ", s).strip()
    return s.upper()


def find_header_index(ws, target_header: str) -> Optional[int]:
    """Return the ZERO-based index of the header that matches target_header (canonical)."""
    target_c = canon(target_header)
    for i, cell in enumerate(ws[1]):  # header row
        if canon(cell.value) == target_c:
            return i
    return None


def worksheet_headers(ws) -> List[str]:
    return [cell.value for cell in ws[1]]


def get_row_dict(ws, row_idx_1based: int) -> Dict[str, Any]:
    headers = worksheet_headers(ws)
    values = [cell.value for cell in ws[row_idx_1based]]
    return {str(h): v for h, v in zip(headers, values)}


# -----------------------------
# SQLite overrides (persisting user-edited values)
# -----------------------------
def db():
    conn = sqlite3.connect(OVERRIDES_DB)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    db_dir = os.path.dirname(OVERRIDES_DB) or "."
    os.makedirs(db_dir, exist_ok=True)
    conn = db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS overrides(
            sheet TEXT NOT NULL,
            client_code TEXT NOT NULL,
            column_canon TEXT NOT NULL,
            column_actual TEXT NOT NULL,
            new_value TEXT NOT NULL,
            updated_at INTEGER NOT NULL,
            PRIMARY KEY (sheet, client_code, column_canon)
        )
    """)
    conn.commit()
    conn.close()


def apply_overrides(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Merge base rows with overrides so user changes persist even if Excel refreshes."""
    if not rows:
        return rows
    conn = db()
    ovs = conn.execute("SELECT sheet, client_code, column_actual, new_value FROM overrides").fetchall()
    conn.close()
    idx = {(r["sheet"], r["client_code"], r["column_actual"]): r["new_value"] for r in ovs}
    out = []
    for r in rows:
        sheet = str(r.get("SOURCE_SHEET", ""))
        code = str(r.get("CLIENT CODE", ""))
        nr = dict(r)
        for k in list(nr.keys()):
            key = (sheet, code, k)
            if key in idx:
                nr[k] = idx[key]
        out.append(nr)
    return out


# -----------------------------
# Excel load / reload (watchdog on file changes)
# -----------------------------
def load_excel():
    global final_df
    try:
        if not os.path.exists(EXCEL_FILE):
            app.logger.warning(f"Excel file not found: {EXCEL_FILE}")
            final_df = pd.DataFrame()
            return

        combined_data = []
        for sheet in SHEETS:
            try:
                df = pd.read_excel(EXCEL_FILE, sheet_name=sheet, engine="openpyxl")
                df["SOURCE_SHEET"] = sheet
                combined_data.append(df)
            except Exception as e:
                app.logger.error(f"Error reading sheet '{sheet}': {e}")

        final_df = pd.concat(combined_data, ignore_index=True) if combined_data else pd.DataFrame()
        app.logger.info(f"✅ Excel reloaded from {EXCEL_FILE}. Rows: {final_df.shape[0]}")
    except Exception as e:
        app.logger.error(f"❌ Error loading Excel: {e}")
        final_df = pd.DataFrame()


class ReloadHandler(FileSystemEventHandler):
    def on_modified(self, event):
        try:
            if os.path.basename(event.src_path) == os.path.basename(EXCEL_FILE):
                time.sleep(0.5)  # avoid partial write race
                load_excel()
        except Exception:
            app.logger.warning("Watchdog handler error; continuing.", exc_info=True)


def start_watcher():
    try:
        event_handler = ReloadHandler()
        observer = Observer()
        watch_dir = os.path.dirname(os.path.abspath(EXCEL_FILE)) or "."
        observer.schedule(event_handler, watch_dir, recursive=False)
        observer.start()
        app.logger.info(f"🔎 Watching for changes in: {watch_dir}")
    except Exception:
        app.logger.warning("Watchdog not started (unsupported environment).", exc_info=True)


def _initialize_once():
    """Run initialization only once per process (safe for Gunicorn workers)."""
    global _initialized
    with _init_lock:
        if _initialized:
            return
        app.logger.info(f"EXCEL_FILE: {EXCEL_FILE} (exists={os.path.exists(EXCEL_FILE)})")
        app.logger.info(f"OVERRIDES_DB: {OVERRIDES_DB}")
        app.logger.info(f"SHEETS: {SHEETS}")
        init_db()
        load_excel()
        # Watchdog is best-effort; ok if it can't run on the platform
        try:
            threading.Thread(target=start_watcher, daemon=True).start()
        except Exception:
            app.logger.exception("Failed to start file watcher")
        _initialized = True


# -----------------------------
# Initialize at import time (Flask 3.x compatible)
# -----------------------------
_initialize_once()


# -----------------------------
# Routes
# -----------------------------
@app.route("/", methods=["GET"])
def home():
    return "✅ API is running! Use /data to get data or POST /update to update Excel."


@app.route("/health", methods=["GET"])
def health():
    exists = os.path.exists(EXCEL_FILE)
    return jsonify({
        "status": "ok",
        "excel_file": EXCEL_FILE,
        "excel_exists": bool(exists),
        "sheets": SHEETS,
        "rows_cached": 0 if final_df is None else int(final_df.shape[0])
    })


@app.route("/data", methods=["GET"])
def get_all_data():
    base = final_df.to_dict(orient="records") if not final_df.empty else []
    merged = apply_overrides(base)
    return jsonify(merged)


@app.route("/update", methods=["POST"])
def update_excel():
    """
    JSON:
    {
      "sheet": "corp",
      "client_code": "C001",
      "column": "Status",   # (visible header; punctuation/case tolerant)
      "new_value": "Shared Client"
    }
    """
    try:
        data = request.json or {}
        sheet = data.get("sheet")
        client_code = (data.get("client_code") or "").strip()
        column_visible = data.get("column")
        new_value = data.get("new_value")

        if not all([sheet, client_code, column_visible]):
            return jsonify({"status": "error", "message": "Missing sheet, client_code, or column"}), 400

        # Optional strict validation for status
        if canon(column_visible) == "STATUS":
            if canon(str(new_value)) not in ALLOWED_STATUS_VALUES:
                return jsonify({"status": "error",
                                "message": "Invalid status. Use 'Cross-Sell' or 'Shared Client'."}), 400

        if not os.path.exists(EXCEL_FILE):
            return jsonify({"status": "error",
                            "message": f"Excel file not found at {EXCEL_FILE}. Upload it to the persistent disk."}), 500

        wb = load_workbook(EXCEL_FILE)
        if sheet not in wb.sheetnames:
            return jsonify({"status": "error", "message": f"Sheet '{sheet}' not found"}), 404
        ws = wb[sheet]

        # Header resolution
        client_code_col_idx = find_header_index(ws, "CLIENT CODE")
        if client_code_col_idx is None:
            return jsonify({"status": "error", "message": "CLIENT CODE header not found"}), 400

        target_col_idx = find_header_index(ws, column_visible)
        if target_col_idx is None:
            return jsonify({"status": "error", "message": f"Column '{column_visible}' not found"}), 400

        # Find the row by client code (case-insensitive)
        target_row_idx = None
        for row in ws.iter_rows(min_row=2):
            cell_val = row[client_code_col_idx].value
            if cell_val is not None and str(cell_val).strip().lower() == client_code.lower():
                target_row_idx = row[0].row  # 1-based
                break

        if target_row_idx is None:
            return jsonify({"status": "error", "message": f"Client Code '{client_code}' not found"}), 404

        # Write new value
        ws.cell(row=target_row_idx, column=target_col_idx + 1, value=new_value)

        # Timestamp column
        ts_header_idx = find_header_index(ws, "STATUS_UPDATED_AT")
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if ts_header_idx is None:
            new_ts_col_1based = ws.max_column + 1
            ws.cell(row=1, column=new_ts_col_1based, value="STATUS_UPDATED_AT")
            ws.cell(row=target_row_idx, column=new_ts_col_1based, value=current_time)
        else:
            ws.cell(row=target_row_idx, column=ts_header_idx + 1, value=current_time)

        # Save Excel (handle locks)
        try:
            wb.save(EXCEL_FILE)
        except PermissionError:
            return jsonify({"status": "error",
                            "message": "Excel file is open/locked. Please close it and retry."}), 500

        # Persist override (bulletproof against ETL/refresh)
        headers = worksheet_headers(ws)
        actual_header = headers[target_col_idx] if target_col_idx < len(headers) else column_visible
        now_epoch = int(time.time())

        db_dir = os.path.dirname(OVERRIDES_DB) or "."
        os.makedirs(db_dir, exist_ok=True)
        conn = db()
        conn.execute("""
            INSERT INTO overrides(sheet, client_code, column_canon, column_actual, new_value, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(sheet, client_code, column_canon)
            DO UPDATE SET new_value=excluded.new_value,
                          updated_at=excluded.updated_at,
                          column_actual=excluded.column_actual
        """, (sheet, client_code, canon(column_visible), actual_header, str(new_value), now_epoch))
        conn.commit()
        conn.close()

        # Small delay & reload cache
        time.sleep(0.3)
        load_excel()

        # Confirm by reading back live value
        live_row = get_row_dict(ws, target_row_idx)
        return jsonify({
            "status": "success",
            "message": f"Updated {actual_header} for {client_code} to '{new_value}' at {current_time}",
            "sheet": sheet,
            "client_code": client_code,
            "column_actual": actual_header,
            "new_value": live_row.get(actual_header, new_value),
            "updated_at": current_time
        })
    except Exception as e:
        app.logger.error("Error:\n" + traceback.format_exc())
        return jsonify({"status": "error", "message": str(e)}), 500


# -----------------------------
# Boot (local run)
# -----------------------------
if __name__ == "__main__":
    _initialize_once()
    port = int(os.environ.get("PORT", "5000"))
    app.run(host="0.0.0.0", port=port, debug=False, use_reloader=False)




